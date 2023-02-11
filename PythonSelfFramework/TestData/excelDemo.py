import openpyxl


book = openpyxl.load_workbook("C:\\Users\\AFisenko\\Excel_file\\Book1.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "Andrey"
print(sheet.cell(row=2, column=2).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A5'].value)

for i in range(1, sheet.max_row + 1):
    print(sheet.cell(row=i, column=1).value)

for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value)

for i in range(1, sheet.max_row + 1): #to get rows
    if sheet.cell(row=i, column=1).value == "Testcase2":

        for j in range(1, sheet.max_column+1): #to get columns
            print(sheet.cell(row=i, column=j).value)

# to start from second row
for i in range(1, sheet.max_row + 1):  # to get rows
    if sheet.cell(row=i, column=1).value == "Testcase2":

        for j in range(2, sheet.max_column + 1):  # to get columns
            print(sheet.cell(row=i, column=j).value)